#!/usr/bin/env python3
"""
Gera planilha Excel com os dados normalizados do CSV para analise.
"""
import csv, re, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT = os.path.join(ROOT, 'venda_aparelhos.csv')
OUTPUT = os.path.join(ROOT, 'scripts', 'vendas_normalizadas.xlsx')

sys_path = os.path.dirname(os.path.abspath(__file__))
import sys
if sys_path not in sys.path:
    sys.path.insert(0, sys_path)

from analisar_importacao import (
    parse_brl, to_date, extrair_troca, normalizar_forma,
    LOJA_MAP, VENDEDOR_MAP, limpar_acentos
)

# ====================================================================
# HELPERS
# ====================================================================

def extrair_estado(modelo):
    m = modelo.upper()
    if 'NOVO' in m and 'SEMINOVO' not in m and 'LACRADO' not in m: return 'novo'
    if 'LACRADO' in m: return 'novo'
    if 'SEMINOVO' in m: return 'seminovo'
    if 'USADO' in m: return 'usado'
    return 'seminovo'

def limpar_modelo(modelo):
    for s in [' NOVO', ' SEMINOVO', ' USADO', ' LACRADO', ' - GARANTIA']:
        if modelo.upper().endswith(s):
            modelo = modelo[:-len(s)]
    return modelo.strip()

def style_header(ws, row=1):
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

# ====================================================================
# LER E PROCESSAR
# ====================================================================

with open(INPUT, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    linhas = list(reader)

print(f'Processando {len(linhas)} linhas...')

vendas_data = []
trocas_data = []
problemas_data = []

for row in linhas:
    data = row.get('DATA', '').strip()
    data_iso = to_date(data)
    modelo_orig = row.get('MODELO', '').strip()
    imei = row.get('IMEI', '').strip().replace(' ', '')
    valor_venda = parse_brl(row.get('VALOR DE VENDA', ''))
    brinde = parse_brl(row.get('BRINDE', ''))
    custo = parse_brl(row.get('CUSTO APARELHO', ''))
    lucro = parse_brl(row.get('LUCRO', ''))
    vendedor = row.get('VENDEDOR', '').strip().title()
    loja = row.get('LOJA', '').strip().upper() or 'CELL'
    forma_orig = row.get('FORMA DE PAGAMENTO', '').strip()

    if valor_venda is None:
        problemas_data.append({
            'modelo': modelo_orig, 'data': data, 'vendedor': vendedor,
            'loja': loja,
            'tipo': 'VENDA NAO MONETARIA', 'detalhe': f'Valor: {row.get("VALOR DE VENDA")} - ignorado'
        })
        continue

    estado = extrair_estado(modelo_orig)
    modelo_limpo = limpar_modelo(modelo_orig)
    formas = normalizar_forma(forma_orig)
    tem_troca = 'troca_aparelho' in formas
    trocas = extrair_troca(forma_orig) if tem_troca else []
    loja_id = LOJA_MAP.get(loja)
    vendedor_id = VENDEDOR_MAP.get(vendedor.upper())
    
    problemas = []
    if not imei: problemas.append('sem_imei')
    if not vendedor_id: problemas.append('sem_vendedor_id')
    
    ven = {
        'data': data,
        'data_iso': data_iso or '',
        'modelo_original': modelo_orig,
        'modelo_limpo': modelo_limpo,
        'imei': imei or '(sem IMEI)',
        'valor_venda': valor_venda,
        'brinde': brinde,
        'custo': custo,
        'lucro': lucro,
        'margem_pct': round((lucro / valor_venda * 100), 1) if valor_venda > 0 else 0,
        'formas': '+'.join(formas),
        'tem_troca': 'SIM' if tem_troca else 'NAO',
        'qtd_trocas': len(trocas),
        'modelo_troca': '; '.join(t['modelo'] for t in trocas) or '',
        'valor_troca': sum(t['valor'] for t in trocas) or 0,
        'vendedor': vendedor,
        'vendedor_id': vendedor_id or '(sem ID)',
        'loja': loja,
        'loja_id': loja_id or '(sem ID)',
        'estado': estado,
        'tipo_pagto_principal': 'pix',
        'problemas': ', '.join(problemas) if problemas else '',
    }
    
    # Payment type detection
    if 'cartao_credito' in formas: ven['tipo_pagto_principal'] = 'cartao_credito'
    elif 'dinheiro' in formas: ven['tipo_pagto_principal'] = 'dinheiro'
    elif 'cartao_debito' in formas: ven['tipo_pagto_principal'] = 'cartao_debito'
    elif tem_troca: ven['tipo_pagto_principal'] = 'troca+outros'
    
    vendas_data.append(ven)
    
    for t in trocas:
        trocas_data.append({
            'venda_data': data,
            'venda_modelo': modelo_orig,
            'venda_valor': valor_venda,
            'vendedor': vendedor,
            'loja': loja,
            'modelo_troca': t['modelo'],
            'valor_troca': t['valor'],
            'forma_orig': forma_orig[:80],
        })
    
    if problemas:
        for p in problemas:
            problemas_data.append({
                'modelo': modelo_orig, 'data': data,
                'vendedor': vendedor, 'loja': loja,
                'tipo': p, 'detalhe': forma_orig[:80],
            })

# ====================================================================
# CRIAR PLANILHA
# ====================================================================

wb = openpyxl.Workbook()

# --- Sheet 1: VENDAS ---
ws1 = wb.active
ws1.title = 'Vendas'

headers_vendas = [
    'DATA', 'DATA ISO', 'MODELO ORIGINAL', 'MODELO LIMPO', 'IMEI',
    'VALOR VENDA', 'BRINDE', 'CUSTO', 'LUCRO', 'MARGEM %',
    'FORMAS PGTO', 'TEM TROCA', 'QTD TROCAS', 'MODELO TROCA', 'VALOR TROCA',
    'VENDEDOR', 'VENDEDOR ID', 'LOJA', 'LOJA ID', 'ESTADO',
    'TIPO PGTO PRINCIPAL', 'PROBLEMAS'
]
ws1.append(headers_vendas)

for v in vendas_data:
    ws1.append([
        v['data'], v['data_iso'], v['modelo_original'], v['modelo_limpo'], v['imei'],
        v['valor_venda'], v['brinde'], v['custo'], v['lucro'], v['margem_pct'],
        v['formas'], v['tem_troca'], v['qtd_trocas'], v['modelo_troca'], v['valor_troca'],
        v['vendedor'], v['vendedor_id'], v['loja'], v['loja_id'], v['estado'],
        v['tipo_pagto_principal'], v['problemas'],
    ])

style_header(ws1)

# Color problem rows
problema_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
for row_idx in range(2, ws1.max_row + 1):
    if ws1.cell(row=row_idx, column=headers_vendas.index('PROBLEMAS') + 1).value:
        for col_idx in range(1, ws1.max_column + 1):
            ws1.cell(row=row_idx, column=col_idx).fill = problema_fill

# Color trade-in rows
troca_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
for row_idx in range(2, ws1.max_row + 1):
    if ws1.cell(row=row_idx, column=headers_vendas.index('TEM TROCA') + 1).value == 'SIM':
        for col_idx in range(1, ws1.max_column + 1):
            fill = ws1.cell(row=row_idx, column=col_idx).fill
            if fill == PatternFill(): # don't overwrite yellow
                ws1.cell(row=row_idx, column=col_idx).fill = troca_fill

auto_width(ws1)

# --- Sheet 2: TROCAS ---
ws2 = wb.create_sheet('Trocas')
headers_trocas = ['DATA VENDA', 'MODELO VENDIDO', 'VALOR VENDA', 'VENDEDOR', 'LOJA',
                  'MODELO TROCA', 'VALOR TROCA', 'FORMA ORIGINAL']
ws2.append(headers_trocas)
for t in trocas_data:
    ws2.append([t['venda_data'], t['venda_modelo'], t['venda_valor'], t['vendedor'],
                t['loja'], t['modelo_troca'], t['valor_troca'], t['forma_orig']])
style_header(ws2)
auto_width(ws2)

# --- Sheet 3: PROBLEMAS ---
ws3 = wb.create_sheet('Problemas')
headers_prob = ['MODELO', 'DATA', 'VENDEDOR', 'LOJA', 'TIPO', 'DETALHE']
ws3.append(headers_prob)
for p in problemas_data:
    ws3.append([p['modelo'], p['data'], p['vendedor'], p['loja'], p['tipo'], p['detalhe']])
style_header(ws3)
auto_width(ws3)

# --- Sheet 4: RESUMO ---
ws4 = wb.create_sheet('Resumo')

def bold(ws, row, col, val, size=12):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=True, size=size)
    return cell

ws4.merge_cells('A1:C1')
bold(ws4, 1, 1, 'RESUMO DA IMPORTACAO', 14)

total_venda = sum(v['valor_venda'] for v in vendas_data)
total_custo = sum(v['custo'] for v in vendas_data)
total_lucro = sum(v['lucro'] for v in vendas_data)
total_brinde = sum(v['brinde'] for v in vendas_data)

resumo_items = [
    ('Registros processados', '', len(vendas_data)),
    ('Total VENDA', f'R$ {total_venda:,.2f}', ''),
    ('Total CUSTO', f'R$ {total_custo:,.2f}', ''),
    ('Total BRINDE', f'R$ {total_brinde:,.2f}', ''),
    ('Total LUCRO', f'R$ {total_lucro:,.2f}', ''),
    ('Margem media', f'{(total_lucro/total_venda*100):.1f}%', ''),
    ('', '', ''),
    ('Vendas com troca', f'{len(trocas_data)} aparelhos', ''),
    ('Valor total trocas', f'R$ {sum(t["valor_troca"] for t in trocas_data):,.2f}', ''),
    ('', '', ''),
    ('Vendedores sem ID', f'{len([p for p in problemas_data if p["tipo"] == "sem_vendedor_id"])}', '(Angel)'),
    ('Sem IMEI', f'{len([p for p in problemas_data if p["tipo"] == "sem_imei"])}', ''),
]

for i, (label, val1, val2) in enumerate(resumo_items):
    row = i + 3
    ws4.cell(row=row, column=1, value=label)
    ws4.cell(row=row, column=2, value=val1)
    if val2:
        ws4.cell(row=row, column=3, value=val2)

# Por loja
row_offset = len(resumo_items) + 5
bold(ws4, row_offset, 1, 'POR LOJA', 12)
ws4.append([])  # skip
ws4.append(['LOJA', 'LOJA ID', 'QTD', 'TOTAL VENDA', 'TOTAL LUCRO', 'TOTAL TROCAS'])
style_header(ws4, row_offset + 2)

lojas_agg = defaultdict(lambda: {'qtd': 0, 'venda': 0, 'lucro': 0, 'trocas': 0})
for v in vendas_data:
    l = v['loja']
    lojas_agg[l]['qtd'] += 1
    lojas_agg[l]['venda'] += v['valor_venda']
    lojas_agg[l]['lucro'] += v['lucro']
    if v['tem_troca'] == 'SIM':
        lojas_agg[l]['trocas'] += v['qtd_trocas']

for loja, agg in sorted(lojas_agg.items()):
    ws4.append([loja, LOJA_MAP.get(loja), agg['qtd'],
                f"R$ {agg['venda']:,.2f}", f"R$ {agg['lucro']:,.2f}", agg['trocas']])

# Por vendedor
row_offset_v = row_offset + 2 + len(lojas_agg) + 3
bold(ws4, row_offset_v, 1, 'POR VENDEDOR', 12)
ws4.append([])
ws4.append(['VENDEDOR', 'VENDEDOR ID', 'QTD', 'TOTAL VENDA', 'TOTAL LUCRO'])
style_header(ws4, row_offset_v + 2)

vend_agg = defaultdict(lambda: {'qtd': 0, 'venda': 0, 'lucro': 0})
for v in vendas_data:
    vn = v['vendedor']
    vend_agg[vn]['qtd'] += 1
    vend_agg[vn]['venda'] += v['valor_venda']
    vend_agg[vn]['lucro'] += v['lucro']

for vendedor, agg in sorted(vend_agg.items()):
    vid = VENDEDOR_MAP.get(vendedor.upper(), '')
    ws4.append([vendedor, vid, agg['qtd'],
                f"R$ {agg['venda']:,.2f}", f"R$ {agg['lucro']:,.2f}"])

auto_width(ws4)

# ====================================================================
# SALVAR
# ====================================================================
wb.save(OUTPUT)
print(f'\nPlanilha salva: {OUTPUT}')
print(f'  Sheet "Vendas":     {len(vendas_data)} linhas')
print(f'  Sheet "Trocas":     {len(trocas_data)} linhas')
print(f'  Sheet "Problemas":  {len(problemas_data)} linhas')
print(f'  Sheet "Resumo":     resumo agregado')
